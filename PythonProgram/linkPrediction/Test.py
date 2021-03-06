'''
@Time: 2019/12/25 19:29
@Author: mih
@Des: 
'''
import scipy.sparse as sp
import numpy
import networkx
import matplotlib.pyplot as plt
from torch import nn
from copy import copy
from Tools import get_test_matrix
import torch
from GraphEmbedding_RandomWalk.NN import LineNetwork
import openpyxl
import math
from openpyxl.styles import PatternFill
from matplotlib.pyplot import MultipleLocator
from sklearn.metrics import roc_auc_score
from torch import nn

# 定义一张用于测试的图, A表示改图的邻接矩阵
A = numpy.array([
    [0, 1, 1, 1, 0, 0, 0],
    [1, 0, 1, 0, 0, 0, 1],
    [1, 1, 0, 0, 1, 0, 1],
    [1, 0, 0, 0, 1, 1, 0],
    [0, 0, 1, 1, 0, 1, 0],
    [0, 0, 0, 1, 1, 0, 1],
    [0, 1, 1, 0, 0, 1, 0]
])
# 根据图的邻接矩阵绘制改图
def draw_graph(A):
    # 节点数
    N = len(A)
    nodes = [i + 1 for i in range(N)]
    edges = []
    for i in range(N):
        j = i
        while(j < N):
            if A[i][j] == 1:
                edges.append((i + 1, j + 1))
            j = j + 1
    # 绘制无向图
    G = networkx.Graph()
    edges = G.edges
    G.add_nodes_from(nodes)
    G.add_edges_from(edges)
    # for edge in G.edges:
    #     print(edge)
    # networkx.draw(G, with_labels=True)
    # plt.show()
    A = networkx.adjacency_matrix(G).todense()
    neighbors = [i-1 for i in G.neighbors(1)]
    sub_A = A[neighbors, :][:, neighbors]
    print(neighbors)
    print(sub_A)
def Test3():
    G = networkx.Graph()
    nodes = numpy.arange(start = 1, stop = 6)
    print("nodes: {0}".format(nodes))
    edges = numpy.array([
        [1, 2], [1, 3], [2, 3], [2, 4], [3, 4], [4, 5]
    ])
    print("edges: {0}".format(edges))
    G.add_edges_from(edges)
    A = networkx.adjacency_matrix(G).todense()
    A = torch.tensor(A, dtype = torch.float)
    print("A: {0}".format(A))
    A_sum = torch.sum(A, dim = 1).reshape([-1, ])
    D = torch.diag(A_sum)
    print("A_sum: {0}".format(A_sum))
    print("D: {0}".format(D))
    L = D - A
    eval, evecs = torch.eig(L.t(), eigenvectors=True)
    print("eval: {0}".format(eval))
    print("evecs: {0}".format(evecs))

def Test3_():
    G = networkx.Graph()
    nodes = numpy.arange(start=1, stop=6)
    print("nodes: {0}".format(nodes))
    edges = numpy.array([
        [1, 2], [1, 3], [2, 3], [2, 4], [3, 4], [4, 5]
    ])
    print("edges: {0}".format(edges))
    G.add_edges_from(edges)
    A = networkx.adjacency_matrix(G).todense()
    A = numpy.array(A)
    print("A type: {0}".format(type(A)))
    print("A: {0}".format(A))
    A_sum = numpy.sum(A, axis = 1).reshape([-1, ])
    D = numpy.diag(A_sum)
    print("D: {0}".format(D))
    L = D - A
    eval, evecs = numpy.linalg.eig(L)
    print('特征值：{0}'.format(eval))
    print('特征向量：{0}'.format(evecs))
def Test1():
    A_ = sp.csr_matrix(A, dtype=int)
    print(A_)
def Test2():
    A_ = numpy.array(A).flatten()
    print(A_)


def Test4():
    c_A = copy(A)
    c_A[0][0] = 1
    print(A)
    print(c_A)
    file = r"C:\Users\mihao\Desktop\米昊的东西\result_matrix.txt"
    with open(file, 'w') as file:
        for row in A:
            file.write(str(row))
            file.write('\n')

def Test5():
    _A = get_test_matrix(A, 0.8)
    print(_A)
# 随机生成一幅图
def Test6(N, density):
    nodes = [i for i in range(N)]
    edges = []
    for src in range(N):
        for dst in range(src + 1, N):
            randomValue = numpy.random.random()
            if (randomValue <= density):
                edges.append([src, dst])
    G = networkx.Graph()
    G.add_nodes_from(nodes)
    G.add_edges_from(edges)
    A = networkx.adjacency_matrix(G)
    print(A.todense())
    print(A.diagonal())

def Test7():
    A = numpy.array(range(9)).reshape([3, 3])
    print("A:{0}".format(A))

    D = numpy.diag([numpy.sum(row) for row in A])
    print("D:{0}".format(D))

    A = torch.tensor(A, dtype = torch.float)
    D = torch.tensor(D, dtype = torch.float)

    reuslt1 = A * D
    result2 = torch.mul(A, D)
    result3 = torch.mm(A, D)
    result4 = torch.matmul(A, D)
    print("result1: {0}".format(reuslt1))
    print("result2: {0}".format(result2))
    print("result3: {0}".format(result3))
    print("result4: {0}".format(result4))

def Test8():
    ln = LineNetwork(6, 6, 6)
    file = r'C:\Users\mihao\Desktop\米昊的东西\train\ln.pkl'
    torch.save(ln.state_dict(), file)

def Test9():

    a = numpy.arange(1, 11).reshape([2, 5])
    b = numpy.arange(10, 20).reshape([2, 5])
    s = [a, b]
    print("a:{0}".format(a))
    print("b:{0}".format(b))
    print(numpy.sum(s, axis = 0))
    print(numpy.sum(s, axis = 1))
def Test10():
    a = numpy.random.randn(3, 3)
    a = round(a, 2)
    print(a)


# https://www.cnblogs.com/wanglle/p/11455758.html
#
def Test11():

    numbers = numpy.random.randn(3, 3)
    red_fill = PatternFill("solid", fgColor="FF0000")
    excel_file = openpyxl.load_workbook(r'C:\Users\mihao\Desktop\米昊的东西\result\show.xlsx')
    if (excel_file['A_star1'] == None):
        excel_file.create_sheet("A_star1")
    A_star1_sheet = excel_file['A_star1']

    row_start = 1
    col_start = 1
    for col in range(1, 3 + 1):
        A_star1_sheet.cell(row = row_start, column = col + 1, value = col)
    for row in range(1, 3 + 1):
        A_star1_sheet.cell(row = row + 1, column = col_start, value = row)

    for row in range(3):
        for col in range(3):
            value = numbers[row][col]
            if value <= 0:
                A_star1_sheet.cell(row = row + 2, column = col + 2).fill = red_fill

            A_star1_sheet.cell(row = row + 2, column = col + 2, value = numbers[row][col])

    # 做了修改后要保存
    excel_file.save(r'C:\Users\mihao\Desktop\米昊的东西\result\show.xlsx')
    excel_file.close()

# 展示一维卷积的过程
# pytorch中一维tensor为（batch_size, channels, width）
# 卷积的过程和二维一样
def Test12():
    a = numpy.arange(start = 0, stop = 40)
    a = a.reshape([2, 4, 5])
    a = torch.tensor(a, dtype = torch.float)
    cov1d = torch.nn.Conv1d(in_channels = 4, out_channels = 2, kernel_size = 1)

    pool = torch.nn.MaxPool1d(kernel_size = 3, stride = 2)
    b = cov1d(a)
    print(b)
    print(b.shape)
    b = pool(b)
    print(b)
    print(b.shape)

def Test13():
    b = 2
    F_ = 4
    input = numpy.arange(0, 8).reshape([b, F_]);
    input = torch.tensor(input, dtype = torch.float)
    print("input:\n {0}".format(input))
    input = input.permute([1, 0])
    print("after permute input:\n {0}".format(input))
    input1 = input.repeat(1, b)
    input2 = input1.view(b * b, -1)
    input3 = input.repeat(b, 1)

    print("input1:\n {0}".format(input1))
    print("input2:\n {0}".format(input2))
    print("input3:\n {0}".format(input3))

# unsqueeze(index) 会在tensor的index这个维度上增加一个维度， 大小为1；
# squeeze(index) 会去掉tensor的index这维度上大小为1的维度， 如果大小不为1， 不会发生变化；
def Test14():
    input = numpy.arange(0, 8).reshape([2, 4])
    input = torch.tensor(input, dtype = torch.float)
    print("input:\n{0}".format(input))
    input = input.unsqueeze(1)
    print("input:\n{0}".format(input))
    input = input.unsqueeze(1)
    print("input:\n{0}".format(input))
    input = input.squeeze(1)
    print("input:\n{0}".format(input))

def Test15():
    a = numpy.array([
        [1, 2],
        [2, 3],
        [3, 4]
    ])
    a = torch.tensor(a, dtype = torch.float)
    values, index = torch.max(a, dim = 1)
    print("values:{0}".format(values))
    print("index:{0}".format(index))

def Test16():
    a = numpy.arange(0, 8).reshape([4, 2])
    a = torch.tensor(a, dtype = torch.float)
    print(a)
    b = torch.ones_like(a)

    print(b)

# 对tensor的第i个维度进行排序并返回索引
# 对tensor的第i个维度进行排序，返回top_k， 并返回其索引
def Test17():
    a = numpy.array([
        [2, 4, 5, 1, 7],
        [2, 8, 1, 3, 4]
    ])
    a = torch.tensor(a, dtype = torch.float)
    a_sort, indexs_sort = a.sort(dim = 1, descending = True)
    a_top3, indexs_top3 = torch.topk(a, k = 3, dim = 1)
    print("a_sort:{0}".format(a_sort))
    print("indexs_sort:{0}".format(indexs_sort))

    print("a_top3:{0}".format(a_top3))
    print("indexs_top3:{0}".format(indexs_top3))

def Test18():

    a = numpy.random.randn(3, 3, 3)
    print("a:\n {0}".format(a))
    a = torch.tensor(a, dtype=torch.float)
    index_X, index_Y, index_Z = torch.where(a > 0)
    print("index_X:\n {0}".format(index_X))
    print("index_Y:\n {0}".format(index_Y))
    print("index_Z:\n {0}".format(index_Z))
    # index_X = index_X.reshape([-1, 1])
    # index_Y = index_Y.reshape([-1, 1])
    # index = torch.cat([index_X, index_Y], dim = 1)
    # print("index:\n {0}".format(index))

def Test19():
    a = numpy.arange(0, 18).reshape([3, 2, 3])
    a = torch.tensor(a, dtype = torch.float)
    print(a)
    b = numpy.arange(0, 6).reshape([3, 2])
    b = torch.tensor(b, dtype = torch.float)
    print(b)
    c = torch.mul(a, b)
    print(c)

def Test20():
    a = [1, 2]
    weight = [10, 1]
    label = [1]
    label_tensor = torch.tensor(label, dtype = torch.long)
    weight_tensor = torch.tensor(weight, dtype = torch.float)
    a_tensor = torch.tensor(a, dtype = torch.float).unsqueeze(dim = 0)
    # soft_max = nn.Softmax(dim = -1)
    cross_entropy = nn.CrossEntropyLoss(weight = weight_tensor, reduction = 'sum')
    cross_entropy_result = cross_entropy(a_tensor, target = label_tensor)
    print("cross_entropy_result : {0}".format(cross_entropy_result))

def Test21():
    # X = [
    #     [1, 2],
    #     [3, 4]
    # ]
    # file = r"C:\Users\mihao\Desktop\test.txt"
    # numpy.savetxt(fname = file, X = X, fmt = '%d')
    file = r".\Data\USAir.gml"
    G = networkx.read_gml(file)
    edges = G.edges
    print(edges)

def Test22():
    X_names = ["A", "B"]
    Y_values = [95.1, 96.2]
    Y_values_2 = [90.2, 91.5]
    Y_values_3 = [93.2, 96.1]
    index = numpy.arange(len(X_names))
    plt.bar(x = index - 0.2, height = Y_values, width = 0.2, fc = 'b', label = 'b')
    plt.bar(x=index, height=Y_values_2, width = 0.2, fc='g', label = 'g')
    plt.bar(x=index + 0.2, height=Y_values_3, width = 0.2, fc='r', label = 'r')
    plt.legend()
    # 设置Y轴的刻度范围【90， 100】
    plt.ylim(90, 100)
    # 设置Y轴的最小刻度1
    y_major_locator = MultipleLocator(1)
    ax = plt.gca()
    ax.yaxis.set_major_locator(y_major_locator)
    plt.legend()
    x = range(len(X_names))
    plt.xticks(x, ["A", "B"])
    plt.show()


def Test23():
    adj_matrix = numpy.array(
        [
            [0, 1, 1, 0, 0, 0, 0],
            [1, 0, 0, 1, 1, 0, 0],
            [1, 0, 0, 0, 0, 1, 1],
            [0, 1, 0, 0, 0, 0, 0],
            [0, 1, 0, 0, 0, 0, 0],
            [0, 0, 1, 0, 0, 0, 0],
            [0, 0, 1, 0, 0, 0, 0]
        ]

    )
    A = adj_matrix
    adj_matrix = numpy.matmul(adj_matrix, A)
    print(adj_matrix)

def Test24():
    criterion = nn.CrossEntropyLoss()
    output = torch.randn(3, 5, requires_grad=True)
    label = torch.empty(3, dtype=torch.long).random_(5)
    loss = criterion(output, label)

    print("网络输出为3个5类:")
    print(output)
    print("要计算loss的类别:")
    print(label)
    print("计算loss的结果:")
    print(loss)

    first = [0, 0, 0]
    for i in range(3):
        first[i] = -output[i][label[i]]
    second = [0, 0, 0]
    for i in range(3):
        for j in range(5):
            second[i] += math.exp(output[i][j])
    res = 0
    for i in range(3):
        res += (first[i] + math.log(second[i]))
    print("自己的计算结果：")
    print(res / 3)

def Test25():
    A = numpy.array([
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9]
    ])
    A = torch.tensor(A)
    index = numpy.array([
        [1, 1],
        [2, 2]
        ]
    )
    index = torch.tensor(index, dtype = torch.long)

    print(A)

def Test26():
    y_true = numpy.array([0, 0, 1, 1])
    y_scores = numpy.array([0.1, 0.4, 0.35, 0.8])
    auc = roc_auc_score(y_true, y_scores)
    print(auc)

def mihCrossEntropy(inputs, labels, weight):
    result = 0
    for index, input in enumerate(inputs):
        input = math.e ** input
        y_ = input[labels[index]]
        sumVal = sum(input)
        corss_entropy = y_ / sumVal
        result = -math.log(corss_entropy, math.e) * weight[labels[index]] + result
    print(result / sum(weight))

def Test27():
    inputs = numpy.array([
        [0.3, 0.7],
        [0.6, 0.4]
    ])
    labels = numpy.array([1, 0])
    weight = numpy.array([1, 2])
    mihCrossEntropy(inputs, labels, weight)
    inputs = torch.tensor(inputs, dtype=torch.float)
    labels = torch.tensor(labels, dtype=torch.long)
    weight = torch.tensor(weight, dtype=torch.float)
    cross_entropy = nn.CrossEntropyLoss(weight=weight)
    loss = cross_entropy(inputs, labels)
    print(loss)

def Test28():
    labels = numpy.array([1, 0])
    labels = torch.tensor(labels, dtype=torch.long)
    one_hot = nn.functional.one_hot(labels, 2)
    print(one_hot)

def Test29():
    a = numpy.array([
        [-1, 2, 3],
        [2, -3, 4]
    ])
    a = torch.tensor(a)
    x = numpy.array([1, 2])
    # y = numpy.array([0, 1])
    # x = torch.tensor(x, dtype = torch.long)
    # y = torch.tensor(y, dtype = torch.long)
    # a = a[x, y]
    x = x.reshape((a.shape[0], -1))
    x = x.repeat(a.shape[-1], axis = -1)
    a = a - x

    print(a)

if __name__ == '__main__':
    Test29()
