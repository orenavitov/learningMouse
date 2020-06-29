'''
@Time: 2019/11/14 14:29
@Author: mih
@Des: 
'''
import numpy
from sklearn.metrics import roc_auc_score
import networkx
from copy import deepcopy
import openpyxl
from openpyxl.styles import colors, Font, PatternFill
import matplotlib.pyplot as plt
import copy
import torch.utils.data as Data
import torch
import math

# Auc 计算
def auc(func):
    def process(A):
        # 节点数
        N = A.shape[0]
        # link标签， 存在为1， 不存在为0
        link_label = []
        # link得分
        link_score = []
        Matrix_similarity = func(A)
        for i in range(N):
            for j in range(N):
                if i != j:
                    link_label.append(A[i][j])
                    link_score.append(Matrix_similarity[i][j])
        auc = roc_auc_score(link_label, link_score)
        return auc
    return process

# 准确率计算
def precision(A, Matrix_similarity, train_A, threshold):
    # 节点数
    N = A.shape[0]
    i = 0
    j = i + 1
    T_P = 0
    T_N = 0
    F_P = 0
    F_N = 0
    while (i < N):
        while (j < N):
            if (train_A[i][j] != 1 and Matrix_similarity[i][j] >= threshold):
                if A[i][j] == 1:
                    T_P += 1
                else:
                    T_N += 1
            if (train_A[i][j] != 1 and Matrix_similarity[i][j] < threshold):
                if A[i][j] == 1:
                    F_P += 1
                else:
                    F_N += 1
            if (train_A[i][j] == 1 and Matrix_similarity[i][j] >= threshold):
                T_P += 1
            if (train_A[i][j] == 1 and Matrix_similarity[i][j] < threshold):
                F_P += 1
            j += 1
        i += 1
        j = i + 1
    return T_P, T_N, F_P, F_N

def process_gml_file(file = r'./Data/bio-GE-GT.gml'):
    G = networkx.read_gml(file)
    A = numpy.array(networkx.adjacency_matrix(G).todense())
    edges = G.edges()
    nodes = G.nodes()
    neighbors = {}
    for node in nodes:
        neighbors[node] = G[node]
    return G, A, edges, nodes, neighbors

# 生成GML文件
# src file 为源文件， 源文件格式为边的信息， 如“src dst”
# dstfile 为目标文件的地址， 后缀为.gml
def generate_gml_file(srcfile, dstfile):
    G = networkx.Graph()
    with open(srcfile, "r") as file:
        edges = numpy.loadtxt(file, dtype = int, usecols = (0, 1))
        min = edges.min()
        max = edges.max()
        if (min > 0) :
            edges = edges - min
            max = max - min
        nodes = range(max + 1)
        G.add_edges_from(edges.tolist())
        G.add_nodes_from(nodes)
    networkx.write_gml(G, dstfile)

def find_common_neighbors(G, node1, node2):
    return networkx.common_neighbors(G, node1, node2)

def get_test_matrix(A, keep_radio):
    shape_A = A.shape
    c_A = deepcopy(A)
    _A = numpy.random.random(size=shape_A) - A
    row = shape_A[0]
    col = shape_A[1]
    for i in range(row):
        for j in range(col):
            if i != j and _A[i][j] > (keep_radio - 1):
                c_A[i][j] = 0
    return c_A

# 寻找介数为steps 的邻居节点
def get_steps_neighbor(A, steps):
    neighbors = []
    for step in range(steps):
        A_current = copy.deepcopy(A)
        e = 1
        while (e < (step + 1)):
            A_current = numpy.matmul(A_current, A)
            e = e + 1
        neighbors_step = []
        for row in A_current:
            neighbors_i = []
            for i, w in enumerate(row):
                if w == 1:
                    neighbors_i.append(i)
            neighbors_step.append(neighbors_i)
        neighbors.append(neighbors_step)
    return neighbors


# 将矩阵写入txt
def write_matrix2txt(A, file):
    with open(file, 'w') as file:
        for row in A:
            file.write(str(row))
            file.write('\n')
"""
将矩阵写入excel
A_star 为预测矩阵
A 为真实矩阵
file 为excel文件地址
sheet_name 为excel工作表的名称
"""
def write_matrix2excel(A_star, A, file, sheet_name = 'A'):
    N = A.shape[0]
    red_fill = PatternFill("solid", fgColor="FF0000")
    green_fill = PatternFill("solid", fgColor="B3EE3A")
    excel_file = openpyxl.load_workbook(file)
    sheets = excel_file.sheetnames
    if (sheet_name not in sheets):
        excel_file.create_sheet(sheet_name)
    A_star1_sheet = excel_file[sheet_name]

    row_start = 1
    col_start = 1
    for col in range(1, N + 1):
        A_star1_sheet.cell(row=row_start, column=col + 1, value=col - 1)
    for row in range(1, N + 1):
        A_star1_sheet.cell(row=row + 1, column=col_start, value=row - 1)

    for row in range(N):
        for col in range(N):
            value = A[row][col]
            if value == 1.0:
                A_star1_sheet.cell(row=row + 2, column=col + 2).fill = red_fill
            if value == 0.0:
                A_star1_sheet.cell(row=row + 2, column=col + 2).fill = green_fill

            A_star1_sheet.cell(row=row + 2, column=col + 2, value=A_star[row][col])

    # 做了修改后要保存
    excel_file.save(file)
    excel_file.close()


# 生成一个随机的图, N为节点数量， density为网络密度
def generate_random_graph(N, density):
    nodes = [i for i in range(N)]
    edges = []
    for src in range(N):
        for dst in range(src + 1, N):
            randomValue = numpy.random.random()
            if (randomValue <= density):
                edges.append([src, dst])
    G = networkx.Graph()
    G.add_nodes_from(nodes)
    G.add_edges_from(edges)
    networkx.draw(G, with_labels=True)
    plt.show()
    A = numpy.array(networkx.adjacency_matrix(G).todense())
    return A

# 获得训练集合测试集的DataLoader
# A 为输入的邻接矩阵
# radio 为0.0~1.0之间的float, 表示多少的数据用作训练集
# batch_size 为批训练的数据量
def get_data_loader(A, radio, batch_size = 32, sample_method = 'under_sample', GPU = False):
    node_number = A.shape[0]
    positives = []
    negavives = []
    for i in range(node_number):
        for j in range(i, node_number):
            if i != j:
                label = A[i][j]
                if label == 1:
                    positives.append([i, j, 1])
                if label == 0:
                    negavives.append([i, j, 0])
    edges_size = len(positives)
    none_edges_size = len(negavives)
    if sample_method == 'all_sample':
        numpy.random.shuffle(positives)
        numpy.random.shuffle(negavives)

        train_positives = positives[: int(edges_size * radio)]
        test_positives = positives[int(edges_size * radio):]
        train_negatives = negavives[: int(none_edges_size * radio)]
        test_negatives = negavives[int(none_edges_size * radio):]

        train_positives.extend(train_negatives)
        train_data = train_positives
        test_positives.extend(test_negatives)
        test_data = test_positives
        numpy.random.shuffle(train_data)
        numpy.random.shuffle(test_data)
    if sample_method == 'under_sample':
        numpy.random.shuffle(positives)
        numpy.random.shuffle(negavives)
        train_positives = positives[: int(edges_size * radio)]
        test_positives = positives[int(edges_size * radio):]
        train_negatives = negavives[: int(edges_size * radio)]
        # test_negatives = negavives[int(edges_size * radio):]
        test_negatives = negavives[int(edges_size * radio): edges_size]
        train_positives.extend(train_negatives)
        train_data = train_positives
        test_positives.extend(test_negatives)
        test_data = test_positives
        numpy.random.shuffle(train_data)
        numpy.random.shuffle(test_data)
    if sample_method == 'over_sample':
        pass
    A_test = numpy.zeros(shape=[node_number, node_number])
    for i, index in enumerate(positives):
        row = index[0]
        col = index[1]
        A_test[row][col] = 1
        A_test[col][row] = 1
    train_pairs = [pair[: 2] for pair in train_data]
    train_labels = [pair[-1] for pair in train_data]
    test_pairs = [pair[: 2] for pair in test_data]
    test_labels = [pair[-1] for pair in test_data]
    train_pairs = torch.tensor(train_pairs, dtype=torch.long)
    train_label = torch.tensor(train_labels, dtype=torch.long)
    train_dataSet = Data.TensorDataset(train_pairs, train_label)
    train_loader = Data.DataLoader(
        dataset=train_dataSet,
        batch_size=batch_size,
        shuffle=True,
    )
    test_pairs = torch.tensor(test_pairs, dtype=torch.long)
    test_label = torch.tensor(test_labels, dtype=torch.long)
    test_dataSet = Data.TensorDataset(test_pairs, test_label)
    test_loader = Data.DataLoader(
        dataset=test_dataSet,
        batch_size=batch_size,
        shuffle=True,
    )
    return train_loader, test_loader, A_test

# 计算介数为steps的邻接矩阵和
def Matrix_pre_handle(A, steps, delay):
    N = A.shape[0]
    A_s = []
    # D = numpy.sum(A, axis = 1, keepdims = False)
    # D = numpy.diag(D)
    I = numpy.eye(N)
    for step in range(steps):
        A_current = copy.deepcopy(A)
        e = 1
        while (e < (step + 1)):
            A_current = numpy.matmul(A_current, A)
            e = e + 1
        for i in range(N):
            for j in range(N):
                if (i == j):
                    A_current[i][j] = 0
                # if (A_current[i][j] != 1):
                #     A_current[i][j] = 0
        A_s.append(delay[step] * A_current)
    result = numpy.sum(A_s, axis=0)
    result = result + I
    # result = result + D
    return result

def cal_cos_similary(src, dst):
    result = numpy.matmul(src, dst.T)
    x_1 = (numpy.sum(src ** 2, axis=1) ** 0.5).reshape([-1, 1])

    x_2 = (numpy.sum(src ** 2, axis=1) ** 0.5).reshape([1, -1])
    result = result / numpy.matmul(x_1, x_2)
    return result

if __name__ == '__main__':
    generate_gml_file(r'C:\Users\mihao\Desktop\米昊的东西\USAir.txt',
                      r'C:\Users\mihao\Desktop\米昊的东西\USAir.gml')
